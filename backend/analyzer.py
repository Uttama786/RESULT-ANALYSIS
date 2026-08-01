import os
import re
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def is_valid_subject_code(code: str, name: str = "") -> bool:
    """Verifies that a code is a valid VTU subject code (must contain digits, not be PASS/FAIL legend)."""
    if not code:
        return False
    code_upper = str(code).strip().upper()
    if not any(char.isdigit() for char in code_upper):
        return False
    if code_upper in ["PASS", "FAIL", "ABSENT", "WITHHELD", "REVALUATION", "ABBREVIATION", "NOMENCLATURE"]:
        return False
    if "->" in str(name) or "=>" in str(name):
        return False
    return True

def calculate_pass_fail_stats(total_registered: int, absent_count: int, failed_count: int, passed_count: int = None) -> dict:
    """
    Standard VTU calculation logic using the 'Appeared Students' method.
    Total Registered = Appeared + Absent
    Appeared = Total Registered - Absent
    Passed = Appeared - Failed (if passed_count is None)
    Pass Percentage = (Passed / Appeared) * 100
    Fail Percentage = (Failed / Appeared) * 100
    """
    total_registered = max(0, int(total_registered))
    absent_count = max(0, int(absent_count))
    failed_count = max(0, int(failed_count))

    appeared_count = max(0, total_registered - absent_count)

    if passed_count is None:
        passed_count = max(0, appeared_count - failed_count)
    else:
        passed_count = max(0, int(passed_count))

    pass_percentage = round((passed_count / appeared_count) * 100, 2) if appeared_count > 0 else 0.0
    fail_percentage = round((failed_count / appeared_count) * 100, 2) if appeared_count > 0 else 0.0

    return {
        "total_registered": total_registered,
        "total_students": total_registered,  # alias for backward compatibility
        "absent_count": absent_count,
        "appeared_count": appeared_count,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "pass_percentage": pass_percentage,
        "fail_percentage": fail_percentage
    }

class ResultAnalyzer:
    def __init__(self, student_records: list):
        self.records = student_records

    def analyze(self) -> dict:
        """Performs analytical summaries on student records using the Appeared Students method."""
        if not self.records:
            return {}

        total_students = len(self.records)

        # Categorize records into Absent, Failed, and Passed
        absent_records = []
        failed_records = []
        passed_records = []

        for r in self.records:
            status_u = str(r.get("status", "")).strip().upper()
            sub_results = [
                str(s.get("result", "")).strip().upper()
                for s in r.get("subjects", [])
                if is_valid_subject_code(s.get("code", ""), s.get("name", ""))
            ]

            is_absent = (status_u in ["ABSENT", "AB", "A"]) or (
                len(sub_results) > 0 and all(res in ["A", "AB", "ABSENT"] for res in sub_results)
            )

            if is_absent:
                absent_records.append(r)
            else:
                has_fail_sub = any(res in ["F", "FAIL"] for res in sub_results)
                if "FAIL" in status_u or has_fail_sub:
                    failed_records.append(r)
                else:
                    passed_records.append(r)

        summary_stats = calculate_pass_fail_stats(
            total_registered=total_students,
            absent_count=len(absent_records),
            failed_count=len(failed_records),
            passed_count=len(passed_records)
        )

        # Identify Toppers (excluding completely absent students with 0 marks if appropriate, but reversed sort by total_marks works)
        toppers = sorted(self.records, key=lambda r: r.get("total_marks", 0), reverse=True)
        toppers_list = [{
            "rank": i + 1,
            "usn": t["usn"],
            "name": t["name"],
            "total_marks": t.get("total_marks", 0),
            "max_marks": t.get("max_marks", 0),
            "percentage": t.get("percentage", 0.0),
            "status": t["status"]
        } for i, t in enumerate(toppers[:10])]  # Top 10

        # Identify Failed Students (excluding AB students)
        failed_list = [{
            "usn": f["usn"],
            "name": f["name"],
            "total_marks": f.get("total_marks", 0),
            "percentage": f.get("percentage", 0.0),
            "failed_subjects": [
                sub["code"] for sub in f.get("subjects", [])
                if sub.get("result") in ["F", "FAIL"] and is_valid_subject_code(sub["code"], sub.get("name", ""))
            ]
        } for f in failed_records]

        # Subject-wise analysis
        subject_data = {}
        for r in self.records:
            for sub in r.get("subjects", []):
                code = sub.get("code", "")
                name = sub.get("name", "")

                if not is_valid_subject_code(code, name):
                    continue

                if code not in subject_data:
                    subject_data[code] = {
                        "code": code,
                        "name": name,
                        "registered": 0,
                        "passed": 0,
                        "failed": 0,
                        "absent": 0,
                        "total_marks_sum": 0,
                        "highest_marks": -1,
                        "highest_usn": ""
                    }

                s_info = subject_data[code]
                s_info["registered"] += 1

                res_u = str(sub.get("result", "")).strip().upper()
                if res_u in ["A", "AB", "ABSENT"]:
                    s_info["absent"] += 1
                elif res_u in ["F", "FAIL"]:
                    s_info["failed"] += 1
                elif res_u in ["P", "PASS"]:
                    s_info["passed"] += 1
                else:
                    s_info["passed"] += 1

                tot_m = sub.get("total", 0)
                s_info["total_marks_sum"] += tot_m

                if tot_m > s_info["highest_marks"]:
                    s_info["highest_marks"] = tot_m
                    s_info["highest_usn"] = r.get("usn", "")

        subject_analysis_list = []
        for code, info in subject_data.items():
            s_stats = calculate_pass_fail_stats(
                total_registered=info["registered"],
                absent_count=info["absent"],
                failed_count=info["failed"],
                passed_count=info["passed"]
            )

            appeared = s_stats["appeared_count"]
            avg_marks = round(info["total_marks_sum"] / appeared, 2) if appeared > 0 else 0.0

            subject_analysis_list.append({
                "code": info["code"],
                "name": info["name"],
                "registered": info["registered"],
                "appeared": s_stats["appeared_count"],
                "passed": info["passed"],
                "failed": info["failed"],
                "absent": info["absent"],
                "pass_percentage": s_stats["pass_percentage"],
                "fail_percentage": s_stats["fail_percentage"],
                "average_marks": avg_marks,
                "highest_marks": info["highest_marks"],
                "highest_usn": info["highest_usn"]
            })

        return {
            "summary": summary_stats,
            "toppers": toppers_list,
            "failed_students": failed_list,
            "subject_analysis": subject_analysis_list
        }

    def _detect_semester_and_batch(self, subject_codes) -> tuple[str, str]:
        # Default fallbacks
        semester = "Fourth Semester"
        batch = "Batch 2021-2022"
        
        # 1. Detect semester from subject codes
        # Standard VTU subject codes look like 21CS41 (4th sem), 21MAT41 (4th sem), 18CS61 (6th sem), etc.
        sem_digits = []
        for code in subject_codes:
            match = re.search(r'[A-Za-z]+(\d)', code)
            if match:
                sem_digits.append(int(match.group(1)))
        
        if sem_digits:
            from collections import Counter
            most_common_sem = Counter(sem_digits).most_common(1)[0][0]
            sem_map = {
                1: "First Semester",
                2: "Second Semester",
                3: "Third Semester",
                4: "Fourth Semester",
                5: "Fifth Semester",
                6: "Sixth Semester",
                7: "Seventh Semester",
                8: "Eighth Semester"
            }
            semester = sem_map.get(most_common_sem, f"Semester {most_common_sem}")
            
        # 2. Detect batch from USNs
        years = []
        for r in self.records:
            usn = r.get("usn", "")
            match = re.search(r'(\d{2})[A-Za-z]{2}\d{3}', usn)
            if match:
                years.append(int(match.group(1)))
            else:
                match_any = re.search(r'\d{2}', usn)
                if match_any:
                    years.append(int(match_any.group(0)))
                    
        if years:
            from collections import Counter
            most_common_year = Counter(years).most_common(1)[0][0]
            batch = f"Batch 20{most_common_year:02d}-20{most_common_year+1:02d}"
            
        return semester, batch

    def _detect_department(self, records) -> str:
        # Default fallback
        dept = "CSE"
        
        branches = []
        for r in records:
            usn = r.get("usn", "")
            match = re.search(r'\d{2}([A-Za-z]{2})\d+', usn)
            if match:
                branches.append(match.group(1).upper())
                
        if branches:
            from collections import Counter
            most_common_branch = Counter(branches).most_common(1)[0][0]
            branch_map = {
                "CS": "CSE",
                "IS": "ISE",
                "EC": "ECE",
                "EE": "EEE",
                "ME": "ME",
                "CV": "CIVIL"
            }
            dept = branch_map.get(most_common_branch, most_common_branch)
            
        return dept

    def _build_overall_results_sheet(self, ws):
        # 1. Show grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # 2. Extract unique subject codes in order of appearance
        subjects_dict = {}
        for r in self.records:
            for sub in r["subjects"]:
                code = sub["code"]
                name = sub["name"]
                if is_valid_subject_code(code, name) and code not in subjects_dict:
                    subjects_dict[code] = name
        subject_codes = list(subjects_dict.keys())
        N = len(subject_codes)
        
        # 3. Detect metadata
        semester, batch = self._detect_semester_and_batch(subject_codes)
        dept = self._detect_department(self.records)
        
        # 4. Styling configurations
        title_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
        header_font = Font(name="Segoe UI", size=9, bold=True, color="000000")
        data_font = Font(name="Segoe UI", size=10, color="000000")
        bold_data_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
        
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
        zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") # Subtle zebra striping
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        light_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        def style_cells(row_start, col_start, row_end, col_end, font=None, fill=None, alignment=None, border=None):
            for r in range(row_start, row_end + 1):
                for c in range(col_start, col_end + 1):
                    cell = ws.cell(row=r, column=c)
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment
                    if border is not None:
                        cell.border = border

        # Row 1: Title Block
        ws.row_dimensions[1].height = 25
        title_col_end = 3 + 3 * N + 4 # SL.No, Name, USN + 3*N subjects + TOTAL, PERCENTAGE (%), Remarks, Backlogs
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_col_end)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{semester} Student Result  {batch}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Header heights
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 48  # Taller to show code + name on two lines
        ws.row_dimensions[4].height = 20
        
        # Merge SL.No
        ws.merge_cells("A2:A4")
        ws.cell(row=2, column=1, value="SL.No")
        style_cells(2, 1, 4, 1, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
        
        # Merge Student Name
        ws.merge_cells("B2:B4")
        ws.cell(row=2, column=2, value="Student Name")
        style_cells(2, 2, 4, 2, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
        
        # Merge USN
        ws.merge_cells("C2:C4")
        ws.cell(row=2, column=3, value="USN")
        style_cells(2, 3, 4, 3, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
        
        # Merge Subject Marks Header
        sub_marks_col_end = 3 + 3 * N
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=sub_marks_col_end)
        ws.cell(row=2, column=4, value=f"Subject Marks code wise  {dept} (AFTER REVALUATION)")
        style_cells(2, 4, 2, sub_marks_col_end, header_font, header_fill, Alignment(horizontal="center", vertical="center"), thin_border)
        
        # Merge TOTAL
        tot_col = sub_marks_col_end + 1
        ws.merge_cells(start_row=2, start_column=tot_col, end_row=4, end_column=tot_col)
        ws.cell(row=2, column=tot_col, value="TOTAL")
        style_cells(2, tot_col, 4, tot_col, header_font, header_fill, Alignment(horizontal="center", vertical="center"), thin_border)

        # Merge PERCENTAGE (%)
        pct_col = sub_marks_col_end + 2
        ws.merge_cells(start_row=2, start_column=pct_col, end_row=4, end_column=pct_col)
        ws.cell(row=2, column=pct_col, value="PERCENTAGE (%)")
        style_cells(2, pct_col, 4, pct_col, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
        
        # Merge Remarks
        rem_col = sub_marks_col_end + 3
        ws.merge_cells(start_row=2, start_column=rem_col, end_row=4, end_column=rem_col)
        ws.cell(row=2, column=rem_col, value="Remarks")
        style_cells(2, rem_col, 4, rem_col, header_font, header_fill, Alignment(horizontal="center", vertical="center"), thin_border)
        
        # Merge No. of Backlogs
        back_col = sub_marks_col_end + 4
        ws.merge_cells(start_row=2, start_column=back_col, end_row=4, end_column=back_col)
        ws.cell(row=2, column=back_col, value="No. of Backlogs")
        style_cells(2, back_col, 4, back_col, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
        
        # Row 3 & 4 Subject Details
        for i, code in enumerate(subject_codes):
            col_start = 4 + 3 * i
            col_end = col_start + 2
            ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
            # Show both subject code AND subject name (matching the result PDF)
            subject_name = subjects_dict.get(code, "")
            ws.cell(row=3, column=col_start, value=f"{code}\n{subject_name}")
            style_cells(3, col_start, 3, col_end, header_font, header_fill, Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border)
            
            ws.cell(row=4, column=col_start, value="IM")
            ws.cell(row=4, column=col_start + 1, value="EM")
            ws.cell(row=4, column=col_start + 2, value="T")
            style_cells(4, col_start, 4, col_end, header_font, header_fill, Alignment(horizontal="center", vertical="center"), thin_border)
            
        # Data rows
        for idx, r in enumerate(self.records):
            row_idx = 5 + idx
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (idx % 2 != 0)
            
            # Map subjects to code for easy lookup
            student_sub_map = {}
            backlogs = 0
            for sub in r["subjects"]:
                student_sub_map[sub["code"]] = sub
                res_u = str(sub["result"]).strip().upper()
                if res_u not in ["P", "PASS", "NA", "NOT APPLICABLE"]:
                    backlogs += 1
            
            # SL.No
            ws.cell(row=row_idx, column=1, value=idx + 1)
            style_cells(row_idx, 1, row_idx, 1, data_font, zebra_fill if is_zebra else None, Alignment(horizontal="center"), light_border)
            
            # Student Name
            ws.cell(row=row_idx, column=2, value=r["name"].upper())
            style_cells(row_idx, 2, row_idx, 2, data_font, zebra_fill if is_zebra else None, Alignment(horizontal="left"), light_border)
            
            # USN
            ws.cell(row=row_idx, column=3, value=r["usn"].upper())
            style_cells(row_idx, 3, row_idx, 3, data_font, zebra_fill if is_zebra else None, Alignment(horizontal="center"), light_border)
            
            # Subjects
            for i, code in enumerate(subject_codes):
                c_start = 4 + 3 * i
                sub = student_sub_map.get(code)
                
                if sub:
                    im = sub["internal"]
                    res_code = str(sub["result"]).strip().upper()
                    
                    if res_code == "F":
                        em = sub["external"]  # Display numeric external marks, not "F"
                        failed = True
                        is_special = False
                    elif res_code in ["A", "AB", "ABSENT"]:
                        em = "A"
                        failed = True
                        is_special = True
                    elif res_code in ["W", "WITHHELD"]:
                        em = "W"
                        failed = True
                        is_special = True
                    elif res_code in ["X", "NE", "NOT ELIGIBLE"]:
                        em = "NE"
                        failed = True
                        is_special = True
                    elif res_code in ["NA", "NOT APPLICABLE"]:
                        em = "NA"
                        failed = False
                        is_special = True
                    else:  # P / PASS
                        em = sub["external"]
                        failed = False
                        is_special = False
                    tot = sub["total"]
                else:
                    im = em = tot = ""
                    failed = False
                    is_special = False
                    
                ws.cell(row=row_idx, column=c_start, value=im)
                ws.cell(row=row_idx, column=c_start + 1, value=em)
                ws.cell(row=row_idx, column=c_start + 2, value=tot)
                
                c_fill = zebra_fill if is_zebra else None
                if is_special:
                    style_cells(row_idx, c_start, row_idx, c_start + 2, Font(name="Segoe UI", size=10, bold=True, color="9C6500"), yellow_fill, Alignment(horizontal="center"), light_border)
                elif failed:
                    style_cells(row_idx, c_start, row_idx, c_start + 2, Font(name="Segoe UI", size=10, bold=True, color="9C0006"), red_fill, Alignment(horizontal="center"), light_border)
                else:
                    style_cells(row_idx, c_start, row_idx, c_start + 2, data_font, c_fill, Alignment(horizontal="center"), light_border)
            
            # TOTAL
            ws.cell(row=row_idx, column=tot_col, value=r["total_marks"])
            style_cells(row_idx, tot_col, row_idx, tot_col, bold_data_font, zebra_fill if is_zebra else None, Alignment(horizontal="center"), light_border)

            # PERCENTAGE (%)
            pct_val = f"{r.get('percentage', 0.0):.2f}%"
            ws.cell(row=row_idx, column=pct_col, value=pct_val)
            style_cells(row_idx, pct_col, row_idx, pct_col, bold_data_font, zebra_fill if is_zebra else None, Alignment(horizontal="center"), light_border)
            
            # Remarks
            status_str = r["status"].upper()
            ws.cell(row=row_idx, column=rem_col, value=r["status"])
            
            rem_fill = None
            rem_font = data_font
            if "FAIL" in status_str:
                rem_fill = red_fill
                rem_font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
            elif any(c in status_str for c in ["DISTINCTION", "FIRST CLASS", "SECOND", "PASS"]):
                rem_fill = green_fill
                rem_font = Font(name="Segoe UI", size=10, bold=True, color="006100")
            elif "ABSENT" in status_str:
                rem_fill = yellow_fill
                rem_font = Font(name="Segoe UI", size=10, bold=True, color="9C6500")
                
            style_cells(row_idx, rem_col, row_idx, rem_col, rem_font, rem_fill or (zebra_fill if is_zebra else None), Alignment(horizontal="center"), light_border)
            
            # No. of Backlogs
            ws.cell(row=row_idx, column=back_col, value=backlogs)
            back_font = Font(name="Segoe UI", size=10, bold=True, color="9C0006") if backlogs > 0 else data_font
            back_fill = red_fill if backlogs > 0 else (zebra_fill if is_zebra else None)
            style_cells(row_idx, back_col, row_idx, back_col, back_font, back_fill, Alignment(horizontal="center"), light_border)

        # Add Color Format Key / Legend Box for clear reference by external users
        legend_start_row = 5 + len(self.records) + 2
        
        ws.cell(row=legend_start_row, column=1, value="COLOR FORMAT KEY / LEGEND:")
        ws.cell(row=legend_start_row, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
        
        # Legend Item 1: Soft Green (PASS / DISTINCTION)
        ws.merge_cells(start_row=legend_start_row+1, start_column=1, end_row=legend_start_row+1, end_column=3)
        ws.cell(row=legend_start_row+1, column=1, value="PASS / FIRST CLASS / DISTINCTION")
        style_cells(legend_start_row+1, 1, legend_start_row+1, 3, Font(name="Segoe UI", size=9, bold=True, color="006100"), green_fill, Alignment(horizontal="center", vertical="center"), thin_border)

        # Legend Item 2: Soft Red (FAIL / BACKLOG)
        ws.merge_cells(start_row=legend_start_row+1, start_column=4, end_row=legend_start_row+1, end_column=6)
        ws.cell(row=legend_start_row+1, column=4, value="FAILED SUBJECT / FAIL RESULT / BACKLOG")
        style_cells(legend_start_row+1, 4, legend_start_row+1, 6, Font(name="Segoe UI", size=9, bold=True, color="9C0006"), red_fill, Alignment(horizontal="center", vertical="center"), thin_border)

        # Legend Item 3: Soft Yellow (SPECIAL STATUS)
        ws.merge_cells(start_row=legend_start_row+1, start_column=7, end_row=legend_start_row+1, end_column=12)
        ws.cell(row=legend_start_row+1, column=7, value="SPECIAL STATUS (A: ABSENT | W: WITHHELD | NE: NOT ELIGIBLE | NA: NOT APPLICABLE)")
        style_cells(legend_start_row+1, 7, legend_start_row+1, 12, Font(name="Segoe UI", size=9, bold=True, color="9C6500"), yellow_fill, Alignment(horizontal="center", vertical="center"), thin_border)
            
        # Column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 16
        for i in range(N):
            c_start = 4 + 3 * i
            ws.column_dimensions[get_column_letter(c_start)].width = 6
            ws.column_dimensions[get_column_letter(c_start + 1)].width = 6
            ws.column_dimensions[get_column_letter(c_start + 2)].width = 6
        ws.column_dimensions[get_column_letter(tot_col)].width = 10
        ws.column_dimensions[get_column_letter(pct_col)].width = 16
        ws.column_dimensions[get_column_letter(rem_col)].width = 28
        ws.column_dimensions[get_column_letter(back_col)].width = 15

    def export_to_excel(self, file_path: str):
        """Generates a highly styled Excel workbook using Pandas and OpenPyXL."""
        if not self.records:
            return
            
        analysis = self.analyze()
        summary = analysis["summary"]
        
        # 1. Create DataFrames for other sheets
        # Sheet 2: Subject-wise Analysis
        df_subjects = pd.DataFrame([{
            "Subject Code": s["code"],
            "Subject Name": s["name"],
            "Students Registered": s["registered"],
            "Appeared": s.get("appeared", s["registered"] - s.get("absent", 0)),
            "Passed": s["passed"],
            "Failed": s["failed"],
            "Absent": s.get("absent", 0),
            "Pass Percentage (%)": s["pass_percentage"],
            "Fail Percentage (%)": s.get("fail_percentage", 0.0)
        } for s in analysis["subject_analysis"]])
        
        # Sheet 3: Standings (Toppers & Fails)
        df_toppers = pd.DataFrame([{
            "Rank": t["rank"],
            "USN": t["usn"],
            "Name": t["name"],
            "Score": t["total_marks"],
            "Max Score": t["max_marks"],
            "Percentage (%)": t["percentage"],
            "Class": t["status"]
        } for t in analysis["toppers"]])
        
        df_fails = pd.DataFrame([{
            "USN": f["usn"],
            "Student Name": f["name"],
            "Total Score": f["total_marks"],
            "Failed Subjects Count": len(f["failed_subjects"]),
            "Failed Subjects Codes": ", ".join(f["failed_subjects"])
        } for f in analysis["failed_students"]])

        # 2. Write to Excel using Pandas ExcelWriter with openpyxl engine
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Write a dummy DataFrame for Overall Results sheet which we will fully rebuild below
            pd.DataFrame().to_excel(writer, sheet_name="Overall Results", index=False)
            df_subjects.to_excel(writer, sheet_name="Subject Analysis", index=False)
            df_toppers.to_excel(writer, sheet_name="Class Toppers", index=False)
            df_fails.to_excel(writer, sheet_name="Failed Students", index=False)
            
        # 3. Open workbook and apply custom styling
        wb = openpyxl.load_workbook(file_path)
        
        # Rebuild Overall Results from scratch
        if "Overall Results" in wb.sheetnames:
            del wb["Overall Results"]
        overall_ws = wb.create_sheet("Overall Results", 0)
        self._build_overall_results_sheet(overall_ws)
        
        # Design Styles for other sheets (Modern Slate/Indigo color palette)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        data_font = Font(name="Segoe UI", size=11, color="000000")
        bold_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Slate/Blue
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light grey-blue
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        for name in wb.sheetnames:
            if name == "Overall Results":
                continue
                
            ws = wb[name]
            
            # Show grid lines explicitly
            ws.views.sheetView[0].showGridLines = True
            
            # Format title or metadata inside the sheet if needed
            ws.insert_rows(1, 2)
            ws["A1"] = f"VTU BATCH RESULT ANALYSIS - {name.upper()}"
            ws["A1"].font = title_font
            ws.row_dimensions[1].height = 25
            
            # Readjust all other row dimensions
            ws.row_dimensions[3].height = 24  # Header row height
            
            # Restructure header row formatting
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            # Loop through data rows (starting at row 4)
            for row in range(4, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
                is_zebra = (row % 2 == 0)
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    
                    # Apply zebra striping
                    if is_zebra:
                        cell.fill = zebra_fill
                        
                    # Specific column alignments & styling
                    val = cell.value
                    
                    # Number alignments
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        # Format percentage columns
                        header_val = ws.cell(row=3, column=col).value or ""
                        if "percentage" in header_val.lower() or "%" in header_val:
                            cell.number_format = '0.00"%"'
                        elif "score" in header_val.lower() or "marks" in header_val.lower() or "total" in header_val.lower():
                            cell.number_format = '#,##0'
                            
                    # Alignment and color coding for status strings
                    if isinstance(val, str):
                        val_upper = val.upper()
                        # Pass / Fail highlighting
                        if val_upper in ["FAIL", "F"]:
                            cell.fill = red_fill
                            cell.font = Font(name="Segoe UI", size=11, bold=True, color="C00000")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif val_upper in ["PASS", "P", "PASS CLASS", "FIRST CLASS", "FIRST CLASS WITH DISTINCTION", "SECOND CLASS"]:
                            cell.fill = green_fill
                            cell.font = Font(name="Segoe UI", size=11, bold=True, color="375623")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif val_upper in ["ABSENT", "A"]:
                            cell.fill = yellow_fill
                            cell.font = Font(name="Segoe UI", size=11, bold=True, color="7F6000")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Scan values from row 3 (header) down to find maximum content length
                for cell in col[2:]:  # Skip top blank rows
                    if cell.value:
                        val_len = len(str(cell.value))
                        if isinstance(cell.value, float):
                            val_len += 4
                        max_len = max(max_len, val_len)
                        
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
                
        # 4. Insert Batch Stats Card in the "Overall Results" sheet top-right
        stats_ws = wb["Overall Results"]
        stats_start_col = stats_ws.max_column + 2
        col_let1 = get_column_letter(stats_start_col)
        col_let2 = get_column_letter(stats_start_col + 1)
        
        # Merge cells for statistics title card
        stats_ws.merge_cells(f"{col_let1}1:{col_let2}1")
        stats_ws[f"{col_let1}1"] = "BATCH SUMMARY STATS"
        stats_ws[f"{col_let1}1"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        stats_ws[f"{col_let1}1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Dark Blue
        stats_ws[f"{col_let1}1"].alignment = Alignment(horizontal="center", vertical="center")
        
        stats_items = [
            ("Total Registered", summary["total_students"]),
            ("Appeared Students", summary.get("appeared_count", summary["total_students"] - summary.get("absent_count", 0))),
            ("Passed Students", summary["passed_count"]),
            ("Failed Students", summary["failed_count"]),
            ("Absent Students", summary.get("absent_count", 0)),
            ("Batch Pass %", f"{summary['pass_percentage']}%"),
            ("Batch Fail %", f"{summary.get('fail_percentage', 0.0)}%")
        ]
        
        for i, (label, value) in enumerate(stats_items):
            row_idx = 2 + i
            stats_ws[f"{col_let1}{row_idx}"] = label
            stats_ws[f"{col_let1}{row_idx}"].font = bold_font
            stats_ws[f"{col_let1}{row_idx}"].border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            stats_ws[f"{col_let1}{row_idx}"].fill = zebra_fill
            
            stats_ws[f"{col_let2}{row_idx}"] = value
            stats_ws[f"{col_let2}{row_idx}"].font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
            stats_ws[f"{col_let2}{row_idx}"].border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            stats_ws[f"{col_let2}{row_idx}"].alignment = Alignment(horizontal="right")
            
            if label == "Batch Pass %":
                stats_ws[f"{col_let2}{row_idx}"].fill = green_fill
                
        # Set statistics column widths
        stats_ws.column_dimensions[col_let1].width = 22
        stats_ws.column_dimensions[col_let2].width = 16
        
        # Save modifications
        wb.save(file_path)
        wb.close()
        logger.info(f"Excel report compiled and styled successfully at: {file_path}")
